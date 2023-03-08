# Building charts in excel

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#print(min_column)      1
#print(max_column)      7
#print(min_row)         5
#print(max_row)         7

barchart = BarChart()

# Building references around the table
data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
# Added +1 to mininmum column as the actual data started in column 2, column 1 was headers
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)
# Referencing columns and rows headers around data

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12") # Where we want to add chart

barchart.title = "Sales by Product Line"
barchart.style = 5
wb.save('barchart.xlsx')