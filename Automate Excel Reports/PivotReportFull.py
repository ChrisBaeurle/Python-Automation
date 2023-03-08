# Converting pivot table to excel report, using all automating processes

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

month = 'february'

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# barchart
barchart = BarChart()

# locate data and categories
data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

# adding data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

# make chart
sheet.add_chart(barchart, "B12") # Where we want to add chart
barchart.title = "Sales by Product Line"
barchart.style = 5

# Loop columns with data
for i in range(min_column+1, max_column+1): # (B, G+1)
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = "Currency"

sheet[f'{get_column_letter(min_column)}{max_row+1}'] = 'Total'

sheet['A1'] = "Sales Report"
sheet['A2'] = month
sheet['A1'].font = Font('Arial', bold=True, size=20)
sheet['A2'].font = Font('Arial', bold=True, size=10)

wb.save(f'report_{month}.xlsx')
